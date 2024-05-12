import inspect
import logging
import os
import pytest
import openpyxl
from selenium.webdriver.support import expected_conditions


@pytest.mark.usefixtures("preSetup")
class BaseClass:
    wait = None

    def create_log_file(self):
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        filePath = os.path.join("C:\\Users\\rawat\\PycharmProjects\\Amazon\\Amazon\\Reports", "logfile.log")
        fileHandler = logging.FileHandler(filePath)
        # fileHandler = logging.FileHandler("logfile.log")
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")
        fileHandler.setFormatter(formatter)
        logger.addHandler(fileHandler)  # filehandler object
        logger.setLevel(logging.DEBUG)
        return logger

    def write_to_textFile(self, filename, mode):
        try:
            with open(filename, mode) as f:
                #f.write("Rahul\n1234323432\nRahul12345")
                f.writelines(["Rahul\n", "1234323432\n", "Rahul12345\n"])
        except FileNotFoundError:
            print("File not found")
        except PermissionError:
            print("Permission denied")


    def read_textFile(self, filename, mode):
        with open(filename, mode) as f:
            return f.readlines()

    def element_to_be_clickable(self, element):
        return self.wait.until(expected_conditions.element_to_be_clickable(element))

    def get_multiple_elements(self, elements):
        return self.wait.until(expected_conditions.presence_of_all_elements_located(elements))

    def check_visibility_of_element(self, element):
        return self.wait.until(expected_conditions.visibility_of_element_located(element))

    def read_testdata_from_excel(self):
        dict = {}
        users = []
        workbook = openpyxl.load_workbook("C:\\Users\\rawat\\Downloads\\users.xlsx")
        current_sheet = workbook.active
        count_rows = current_sheet.max_row
        count_column = current_sheet.max_column
        for row in range(2, count_rows + 1):
            for column in range(2, count_column + 1):
                nm = current_sheet.cell(row=1, column=column)
                data = current_sheet.cell(row=row, column=column)
                # print(nm.value, data.value)
                dict[nm.value] = data.value
                users.append(dict)
                dict = {}

        return users
